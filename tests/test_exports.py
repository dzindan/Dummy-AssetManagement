# -*- coding: utf-8 -*-
"""Excel export + import-template verification.

Follows the same isolated-DB pattern as tests/test_auth.py: override
%LOCALAPPDATA% to a fresh temp directory before create_app(), then drive the
app through app.test_client(). Every export/template route sits behind the
login gate (see app/auth.py), so each test logs in an Admin account first.
"""
import io
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from app import create_app  # noqa: E402
from app.auth import create_account  # noqa: E402
from app.db import get_connection  # noqa: E402
from app.importer import (  # noqa: E402
    BRANCH_FILE_REQUIRED_COLUMNS,
    HEADER_ALIASES,
    USER_FILE_REQUIRED_COLUMNS,
    import_asset_report,
    import_branch_file,
    import_user_file,
)


def _fresh_app():
    os.environ["LOCALAPPDATA"] = tempfile.mkdtemp(prefix="am_exporttest_")
    return create_app()


def _role_id(conn, name):
    return conn.execute("SELECT id FROM roles WHERE name = ?", (name,)).fetchone()["id"]


def _wb_from_response(resp):
    return openpyxl.load_workbook(io.BytesIO(resp.get_data()))


def _seed_common(conn):
    """One branch, one asset, one hand-over record, one network-check log
    row, one import-log row - enough for every export route to have exactly
    1 row to find."""
    conn.execute("INSERT INTO branches (branch_no, eng_name) VALUES ('001', 'Hanoi Branch')")
    conn.execute("INSERT INTO import_batches (id, imported_at, kind, period) VALUES (1, datetime('now'), 'asset_report', '2026-01')")
    conn.execute(
        "INSERT INTO asset_items (batch_id, asset_key, branch_no, branch_dept, device_name, device_name_raw, "
        "model_device, serial_tag, status, full_name, user_id_norm, user_id_raw, handover_date) "
        "VALUES (1, 'k1', '001', 'Hanoi', 'PC', 'PC', 'M1', 'SN1', 'IN USE', 'A', '1001', '1001', '2026-01-01')"
    )
    conn.execute(
        "INSERT INTO handover_records (created_at, ho_date, user_no, user_name, branch_no, ho_type, reason, receiving_name) "
        "VALUES (datetime('now'), '2026-01-01', 'U1', 'A', '001', 'PERMANENT', 'test', 'A')"
    )
    conn.execute(
        "INSERT INTO network_check_log (applied_at, branch_no, ip, asset_id, field, old_value, new_value) "
        "VALUES (datetime('now'), '001', '10.0.0.1', 1, 'serial', 'OLD', 'NEW')"
    )
    conn.execute(
        "INSERT INTO import_log (imported_at, kind, source_file, rows_processed, result) "
        "VALUES (datetime('now'), 'asset_report', 'test.xlsx', 1, 'OK')"
    )


class ExportRouteSmokeTests(unittest.TestCase):
    """Every export route (6 new + 4 refactored) returns a parseable .xlsx
    with the expected header row and the expected row count."""

    def setUp(self):
        self.app = _fresh_app()
        conn = get_connection()
        try:
            admin_role_id = _role_id(conn, "Admin")
            _seed_common(conn)
            conn.commit()
        finally:
            conn.close()
        create_account("exporter", "exporterpass1", admin_role_id)
        self.client = self.app.test_client()
        self.client.post("/login", data={"username": "exporter", "password": "exporterpass1"})

    def _assert_valid_xlsx(self, resp, expected_header=None, expected_rows=None):
        self.assertEqual(resp.status_code, 200)
        wb = _wb_from_response(resp)
        data_rows = list(wb.active.iter_rows(values_only=True))
        if expected_header is not None:
            self.assertEqual(data_rows[0], expected_header)
        if expected_rows is not None:
            self.assertEqual(len(data_rows) - 1, expected_rows)
        return data_rows

    def test_dashboard_export(self):
        resp = self.client.get("/export")
        rows = self._assert_valid_xlsx(resp)
        self.assertEqual(rows[0][0], "Branch")

    def test_manage_assets_export(self):
        resp = self.client.get("/assets/export")
        self._assert_valid_xlsx(
            resp,
            expected_header=(
                "Branch", "Device", "Model", "Serial/Service Tag", "Status",
                "Full Name", "User ID", "Position", "Remark", "Handover Date",
            ),
            expected_rows=1,
        )

    def test_branch_detail_export(self):
        resp = self.client.get("/branch/001/export")
        self._assert_valid_xlsx(
            resp,
            expected_header=(
                "Device", "Model", "Serial/Service Tag", "Status", "Full Name",
                "User ID", "Position", "Remark", "Handover Date",
            ),
            expected_rows=1,
        )

    def test_duplicates_export(self):
        conn = get_connection()
        try:
            conn.execute(
                "INSERT INTO asset_items (batch_id, asset_key, branch_no, device_name, device_name_raw, "
                "model_device, serial_tag, status, full_name, user_id_norm, user_id_raw) "
                "VALUES (1, 'k2', '001', 'PC', 'PC', 'M1', 'SN1', 'IN USE', 'B', 'U2', 'U2')"
            )
            conn.commit()
        finally:
            conn.close()
        resp = self.client.get("/assets/duplicates/export")
        self._assert_valid_xlsx(
            resp,
            expected_header=("Serial", "Branch", "Device", "Model", "Full Name", "User ID", "Status", "Asset ID"),
            expected_rows=2,
        )

    def test_import_history_export(self):
        resp = self.client.get("/import/history/export")
        self._assert_valid_xlsx(
            resp,
            expected_header=("Imported At", "Imported By", "Kind", "Source File", "Period", "Rows Processed", "Result"),
            expected_rows=1,
        )

    def test_network_check_log_export(self):
        resp = self.client.get("/network-check/log/export")
        self._assert_valid_xlsx(
            resp,
            expected_header=("Applied At", "Branch", "IP", "Asset ID", "Field", "Old Value", "New Value"),
            expected_rows=1,
        )

    def test_handover_history_export(self):
        resp = self.client.get("/history/export")
        self._assert_valid_xlsx(
            resp,
            expected_header=("Hand-Over Date", "Logged At", "Created By", "User ID", "Branch", "Type", "Reason", "Receiving Party"),
            expected_rows=1,
        )

    def test_user_asset_history_export(self):
        resp = self.client.get("/user-history/export?q=1001")
        self._assert_valid_xlsx(
            resp,
            expected_header=(
                "Period", "Device", "Model", "Serial/Tag", "Status", "Branch", "Handover Date", "Change",
            ),
            expected_rows=1,
        )


class FilteredExportTests(unittest.TestCase):
    """Manage Assets and Hand-Over History exports must respect whatever
    filters are active - export what's on screen, not a separate
    always-everything dump."""

    def setUp(self):
        self.app = _fresh_app()
        conn = get_connection()
        try:
            admin_role_id = _role_id(conn, "Admin")
            conn.execute("INSERT INTO branches (branch_no, eng_name) VALUES ('001', 'Hanoi Branch')")
            conn.execute("INSERT INTO branches (branch_no, eng_name) VALUES ('002', 'HCM Branch')")
            conn.execute("INSERT INTO import_batches (id, imported_at, kind) VALUES (1, datetime('now'), 'asset_report')")
            conn.execute(
                "INSERT INTO asset_items (batch_id, asset_key, branch_no, device_name, device_name_raw, "
                "model_device, serial_tag, status, full_name, user_id_norm, user_id_raw) "
                "VALUES (1, 'k1', '001', 'PC', 'PC', 'M1', 'SN1', 'IN USE', 'A', 'U1', 'U1')"
            )
            conn.execute(
                "INSERT INTO asset_items (batch_id, asset_key, branch_no, device_name, device_name_raw, "
                "model_device, serial_tag, status, full_name, user_id_norm, user_id_raw) "
                "VALUES (1, 'k2', '002', 'LCD', 'LCD', 'M2', 'SN2', 'IN USE', 'B', 'U2', 'U2')"
            )
            conn.execute(
                "INSERT INTO handover_records (created_at, ho_date, user_no, user_name, branch_no, ho_type, reason, receiving_name) "
                "VALUES (datetime('now'), '2026-01-01', 'U1', 'A', '001', 'PERMANENT', 'test', 'A')"
            )
            conn.execute(
                "INSERT INTO handover_records (created_at, ho_date, user_no, user_name, branch_no, ho_type, reason, receiving_name) "
                "VALUES (datetime('now'), '2026-01-02', 'U2', 'B', '002', 'PERMANENT', 'test', 'B')"
            )
            conn.commit()
        finally:
            conn.close()
        create_account("filtertester", "filterpass1", admin_role_id)
        self.client = self.app.test_client()
        self.client.post("/login", data={"username": "filtertester", "password": "filterpass1"})

    def _row_count(self, resp):
        wb = _wb_from_response(resp)
        return len(list(wb.active.iter_rows(values_only=True))) - 1

    def test_manage_assets_export_respects_branch_filter(self):
        unfiltered = self._row_count(self.client.get("/assets/export"))
        filtered = self._row_count(self.client.get("/assets/export?branch_no=001"))
        self.assertEqual(unfiltered, 2)
        self.assertEqual(filtered, 1)

    def test_handover_history_export_respects_branch_filter(self):
        unfiltered = self._row_count(self.client.get("/history/export"))
        filtered = self._row_count(self.client.get("/history/export?branch_no=001"))
        self.assertEqual(unfiltered, 2)
        self.assertEqual(filtered, 1)


class TemplateDownloadTests(unittest.TestCase):
    """The 3 template routes must never drift from what the importer
    actually accepts, and their example row must genuinely be importable -
    not just superficially shaped right."""

    def setUp(self):
        self.app = _fresh_app()
        conn = get_connection()
        try:
            admin_role_id = _role_id(conn, "Admin")
        finally:
            conn.close()
        create_account("tpltester", "tplpass1", admin_role_id)
        self.client = self.app.test_client()
        self.client.post("/login", data={"username": "tpltester", "password": "tplpass1"})

    def test_asset_report_template_header_matches_importer(self):
        resp = self.client.get("/import/templates/asset-report")
        self.assertEqual(resp.status_code, 200)
        wb = _wb_from_response(resp)
        rows = list(wb.active.iter_rows(values_only=True))
        expected = tuple(aliases[0] for aliases in HEADER_ALIASES.values())
        self.assertEqual(rows[0], expected)

    def test_branch_codes_template_header_matches_importer(self):
        resp = self.client.get("/import/templates/branch-codes")
        wb = _wb_from_response(resp)
        rows = list(wb.active.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), BRANCH_FILE_REQUIRED_COLUMNS)

    def test_user_ids_template_header_matches_importer(self):
        resp = self.client.get("/import/templates/user-ids")
        wb = _wb_from_response(resp)
        rows = list(wb.active.iter_rows(values_only=True))
        self.assertEqual(list(rows[0][: len(USER_FILE_REQUIRED_COLUMNS)]), USER_FILE_REQUIRED_COLUMNS)

    def _save_template(self, url, filename) -> str:
        resp = self.client.get(url)
        tmpdir = tempfile.mkdtemp(prefix="am_tplsave_")
        path = os.path.join(tmpdir, filename)
        with open(path, "wb") as f:
            f.write(resp.get_data())
        return path

    def test_asset_report_example_row_actually_imports(self):
        path = self._save_template("/import/templates/asset-report", "asset.xlsx")
        report = import_asset_report(path, source_label="asset.xlsx")
        self.assertEqual(report.error, "")
        self.assertEqual(report.rows_imported, 1)

    def test_branch_codes_example_row_actually_imports(self):
        path = self._save_template("/import/templates/branch-codes", "branch.xlsx")
        result = import_branch_file(path)
        self.assertNotIn("error", result)
        self.assertEqual(result["rows"], 1)

    def test_user_ids_example_row_actually_imports(self):
        path = self._save_template("/import/templates/user-ids", "user.xlsx")
        result = import_user_file(path)
        self.assertNotIn("error", result)
        self.assertEqual(result["rows"], 1)


if __name__ == "__main__":
    unittest.main()
