"""Ingestion & normalization for the two kinds of source spreadsheets:

- ID master files (IDFromAither/*.xlsx): branch master list and user/banker list.
- Asset report files (Asset reports/*.xlsx, and the aggregated Total Asset file):
  messy, inconsistently-shaped equipment lists that need header detection,
  column-alias mapping, branch resolution, and device-name normalization
  before they can be loaded.

Nothing here talks to Flask; routes call into these functions and render the
returned summary objects/dicts.
"""

from __future__ import annotations

import datetime as dt
import json
import re
from dataclasses import dataclass, field

import openpyxl
from unidecode import unidecode

from .db import get_connection, log_import
from .text_utils import normalize_user_id, strip_bank_prefix


def _now_iso() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def _clean_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


PLACEHOLDER_TOKENS = {"NA", "N/A", "NONE", "NULL", "-", "--", "N.A", "N.A."}


def _clean_serial(value) -> str:
    """Like _clean_str, but treats common "no serial recorded" placeholders as
    blank so they don't get flagged as duplicate serials across unrelated assets."""
    text = _clean_str(value)
    return "" if text.upper() in PLACEHOLDER_TOKENS else text



_TO_ABBREVIATION_RE = re.compile(r"\bT\.O\.?\b")


def normalize_branch_text(value) -> str:
    """Uppercased, accent-stripped branch text used everywhere a branch name
    gets compared (alias lookup, resolve_branch's scoring, asset_key). Also
    expands the "T.O" abbreviation (Transaction Office) to its full form, so
    a file labeled e.g. "South Saigon T.O" auto-resolves against a branch
    master eng_name like "...TRANSACTION OFFICE" without needing a manual
    alias first - the space-insensitive matching in resolve_branch already
    absorbs spacing differences like "Saigon" vs "Sai Gon" on top of this."""
    text = unidecode(_clean_str(value)).upper().strip()
    return _TO_ABBREVIATION_RE.sub("TRANSACTION OFFICE", text)


def _normalize_header_cell(value) -> str:
    return re.sub(r"\s+", " ", _clean_str(value)).upper()


# --- Asset-report column detection -----------------------------------------

HEADER_ALIASES = {
    "branch_dept": ["BRANCH / DEPT", "BRANCH/DEPT", "BRANCH DEPT", "BRANCH"],
    "device_name": ["DEVICE NAME", "DEVICE"],
    "user_id": ["USER ID", "USERID"],
    "full_name": ["FULL NAME", "FULLNAME"],
    "model_device": ["MODEL DEVICE", "MODEL"],
    "serial_tag": [
        "SERIAL/ SERVICE TAG",
        "SERIAL / SERVICE TAG",
        "SERIAL/SERVICE TAG",
        "SERIAL NUMBER",
        "SERIAL",
        "SERVICE TAG",
    ],
    "status": ["STATUS"],
    "remark": ["REMARK", "NOTE", "NOTES"],
    "position": ["POSITION", "CURRENT POSITION", "TITLE"],
    "handover_date": ["HANDOVER DATE", "HANDOVERDAY", "HANDOVER DAY", "HANDOVER\nDAY"],
    "ip": ["IP"],
}

# Sheets we should never treat as the equipment list even if a stray header matches.
SHEET_NAME_SKIP_PATTERNS = ["PIVOT", "PIOT", "GUIDELINE", "HDD BROKEN"]

REQUIRED_FIELDS_FOR_HEADER_ROW = {"device_name", "serial_tag"}


def _build_alias_lookup() -> dict[str, str]:
    lookup = {}
    for field_name, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            lookup[_normalize_header_cell(alias)] = field_name
    return lookup


_ALIAS_LOOKUP = _build_alias_lookup()


@dataclass
class HeaderMatch:
    sheet_name: str
    header_row_idx: int  # 0-based
    col_map: dict[str, int]  # field_name -> column index
    branch_hint: str = ""


def _find_branch_hint(ws, header_row_idx: int) -> str:
    """Look for a 'Branch/TO/Center Name:' style label near the top of the sheet."""
    for row in ws.iter_rows(min_row=1, max_row=min(header_row_idx + 1, 10), values_only=True):
        for i, cell in enumerate(row):
            text = _clean_str(cell)
            if text and "BRANCH" in text.upper() and "NAME" in text.upper():
                for after in row[i + 1:]:
                    after_text = _clean_str(after)
                    if after_text:
                        return after_text
    return ""


def detect_equipment_sheet(wb) -> HeaderMatch | None:
    best: HeaderMatch | None = None
    best_score = 0
    for sheet_name in wb.sheetnames:
        upper_name = sheet_name.upper()
        if any(p in upper_name for p in SHEET_NAME_SKIP_PATTERNS):
            continue
        ws = wb[sheet_name]
        max_scan = min(ws.max_row or 0, 10)
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
            col_map: dict[str, int] = {}
            for col_idx, cell in enumerate(row):
                key = _normalize_header_cell(cell)
                if not key:
                    continue
                field_name = _ALIAS_LOOKUP.get(key)
                if field_name and field_name not in col_map:
                    col_map[field_name] = col_idx
            score = len(col_map)
            has_required = REQUIRED_FIELDS_FOR_HEADER_ROW.issubset(col_map.keys())
            if has_required and score > best_score:
                best_score = score
                best = HeaderMatch(
                    sheet_name=sheet_name,
                    header_row_idx=row_idx,
                    col_map=col_map,
                    branch_hint=_find_branch_hint(ws, row_idx),
                )
    return best


def _parse_date(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    text = _clean_str(value)
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text  # keep raw text rather than losing the value


def _asset_key(branch_dept: str, device_name: str, model_device: str, serial_tag: str, user_id_norm: str) -> str:
    serial = _clean_str(serial_tag).upper()
    if serial:
        return f"SN:{serial}"
    return "FB:" + "|".join(
        [
            normalize_branch_text(branch_dept),
            _clean_str(device_name).upper(),
            _clean_str(model_device).upper(),
            user_id_norm,
        ]
    )


# --- Branch resolution --------------------------------------------------

def _branch_match_score(candidate_norm: str, norm: str) -> float:
    """Higher is a better match. Asset reports are about physical branches, so
    when a terse label (e.g. "HANOI") ambiguously substring-matches several
    head-office departments as well as the actual branch, prefer candidates
    that look like a physical location ("BRANCH" / "TRANSACTION OFFICE") over
    departments/divisions/centers/teams, and prefer the closer-length (more
    exact) match among ties."""
    score = 0.0
    if candidate_norm == norm:
        score += 100
    if "BRANCH" in candidate_norm or "TRANSACTION OFFICE" in candidate_norm:
        score += 10
    score -= len(candidate_norm) * 0.01
    return score


def resolve_branch(conn, branch_text: str, cache: dict | None = None) -> tuple[str, str]:
    """Return (branch_no, matched_display_name) for a free-text branch name.

    `cache` lets a caller processing thousands of rows (which usually repeat
    the same handful of branch names) avoid re-scanning the branches table
    for every single row.
    """
    if cache is not None and branch_text in cache:
        return cache[branch_text]
    result = _resolve_branch_uncached(conn, branch_text)
    if cache is not None:
        cache[branch_text] = result
    return result


def _resolve_branch_uncached(conn, branch_text: str) -> tuple[str, str]:
    norm = normalize_branch_text(branch_text)
    if not norm:
        return "", ""

    alias_row = conn.execute(
        "SELECT branch_no FROM branch_aliases WHERE alias = ?", (norm,)
    ).fetchone()
    if alias_row:
        b = conn.execute(
            "SELECT branch_no, eng_name FROM branches WHERE branch_no = ?",
            (alias_row["branch_no"],),
        ).fetchone()
        if b:
            return b["branch_no"], b["eng_name"]

    rows = conn.execute("SELECT branch_no, local_name, eng_name FROM branches").fetchall()

    # Compare with spaces stripped throughout in a single pass - some
    # branches have inconsistent spacing between their own local_name and
    # eng_name (e.g. "HA NOI ..." vs "...HANOI ..."), so a space-preserving
    # pass can match the wrong candidate before a space-insensitive pass
    # ever gets a chance to consider the right one. Scoring every candidate
    # in one pass (see _branch_match_score) avoids that ordering trap.
    norm_nospace = norm.replace(" ", "")
    matches = []  # (score, branch_no, eng_name)
    for r in rows:
        for candidate in (r["local_name"], r["eng_name"]):
            candidate_nospace = normalize_branch_text(candidate).replace(" ", "")
            if not candidate_nospace:
                continue
            if (
                candidate_nospace == norm_nospace
                or norm_nospace in candidate_nospace
                or candidate_nospace in norm_nospace
            ):
                matches.append((_branch_match_score(candidate_nospace, norm_nospace), r["branch_no"], r["eng_name"]))
    if matches:
        matches.sort(key=lambda m: m[0], reverse=True)
        return matches[0][1], matches[0][2]
    return "", ""


def reresolve_unresolved_assets(conn, raw_hint: str, branch_no: str) -> int:
    """After a branch alias is assigned in Settings, retroactively fix any
    already-imported asset_items rows that are still sitting unresolved
    (branch_no = '') for this exact hint - otherwise the Dashboard and
    Manage Assets would keep showing them as "Unresolved" until the next
    time that file happens to be re-imported, which is confusing right
    after you've just fixed the mapping. Matches two ways, since either can
    apply depending on how the row was imported:

    1. The row's own `branch_dept` text matches - true for multi-branch
       imports (Total Asset baseline) where each row is resolved
       independently.
    2. The row's `import_batches.label` matches - true for single-branch
       report files, which stamp *every* row with one file-level
       resolution (see `_ingest_asset_rows`'s `fixed_branch_no`), even when
       a row's own raw `branch_dept` text differs from that file-level
       label (e.g. label "South saigon T.O" but a row's own text just says
       "SOUTH SAIGON").

    Returns how many rows were updated.
    """
    target_norm = normalize_branch_text(raw_hint)
    if not target_norm:
        return 0

    dept_rows = conn.execute(
        "SELECT id, branch_dept FROM asset_items WHERE branch_no = '' AND branch_dept != ''"
    ).fetchall()
    ids = {r["id"] for r in dept_rows if normalize_branch_text(r["branch_dept"]) == target_norm}

    batches = conn.execute("SELECT id, label FROM import_batches").fetchall()
    matching_batch_ids = [b["id"] for b in batches if normalize_branch_text(b["label"]) == target_norm]
    if matching_batch_ids:
        placeholders = ",".join("?" * len(matching_batch_ids))
        batch_rows = conn.execute(
            f"SELECT id FROM asset_items WHERE branch_no = '' AND batch_id IN ({placeholders})",
            matching_batch_ids,
        ).fetchall()
        ids.update(r["id"] for r in batch_rows)

    if ids:
        id_list = list(ids)
        placeholders = ",".join("?" * len(id_list))
        conn.execute(f"UPDATE asset_items SET branch_no = ? WHERE id IN ({placeholders})", [branch_no, *id_list])
    return len(ids)


def record_unresolved_branch(conn, raw_hint: str) -> None:
    """Track a branch label that didn't resolve to any branch, so Settings >
    Branch Name Mapping can surface it for the user to assign - the same
    idea as record_unmapped_device, but for branches. A no-op for blank
    hints (nothing meaningful to show)."""
    raw_hint = _clean_str(raw_hint)
    if not raw_hint:
        return
    now = _now_iso()
    conn.execute(
        """
        INSERT INTO branch_unresolved (raw_hint, first_seen_at, last_seen_at, occurrences)
        VALUES (?, ?, ?, 1)
        ON CONFLICT(raw_hint) DO UPDATE SET
            last_seen_at = excluded.last_seen_at,
            occurrences = occurrences + 1
        """,
        (raw_hint, now, now),
    )


# --- Device/status/model normalization -----------------------------------
# Three fields (device_name, status, model_device) all get the exact same
# treatment: an alias table maps known spelling variants to one canonical
# name, and anything that doesn't match an alias or a standard name passes
# through uppercased/trimmed as-is (never dropped) while getting flagged in
# an "unmapped" table for Settings to surface. Shared here since the three
# only differ in which table they read/write.

def _normalize_via_alias_table(conn, alias_table: str, raw_value, cache: dict | None = None) -> str:
    clean = _clean_str(raw_value).upper()
    if not clean:
        return ""
    if cache is not None and clean in cache:
        return cache[clean]
    row = conn.execute(
        f"SELECT canonical_name FROM {alias_table} WHERE alias = ?", (clean,)
    ).fetchone()
    result = row["canonical_name"] if row else clean
    if cache is not None:
        cache[clean] = result
    return result


def _record_unmapped(conn, table: str, column: str, raw_value: str) -> None:
    now = _now_iso()
    conn.execute(
        f"""
        INSERT INTO {table} ({column}, first_seen_at, last_seen_at, occurrences)
        VALUES (?, ?, ?, 1)
        ON CONFLICT({column}) DO UPDATE SET
            last_seen_at = excluded.last_seen_at,
            occurrences = occurrences + 1
        """,
        (raw_value, now, now),
    )


def normalize_device_name(conn, raw_name: str, cache: dict | None = None) -> str:
    """Map a raw device-name string to its canonical form via device_aliases;
    unmapped names pass through uppercased/trimmed as-is (still usable, just
    flagged elsewhere so it can be mapped later without losing data)."""
    return _normalize_via_alias_table(conn, "device_aliases", raw_name, cache)


def record_unmapped_device(conn, raw_name: str) -> None:
    """Track a device name that isn't yet mapped to any standard name, so the
    Settings > Device Mapping page can surface it for the user to assign."""
    _record_unmapped(conn, "device_unmapped", "raw_name", raw_name)


def normalize_status(conn, raw_status: str, cache: dict | None = None) -> str:
    """Same idea as normalize_device_name, for the Status column."""
    return _normalize_via_alias_table(conn, "status_aliases", raw_status, cache)


def record_unmapped_status(conn, raw_status: str) -> None:
    """Same idea as record_unmapped_device, for the Status column."""
    _record_unmapped(conn, "status_unmapped", "raw_status", raw_status)


def normalize_model_device(conn, raw_model: str, cache: dict | None = None) -> str:
    """Same idea as normalize_device_name, for the Model column."""
    return _normalize_via_alias_table(conn, "model_aliases", raw_model, cache)


def record_unmapped_model(conn, raw_model: str) -> None:
    """Same idea as record_unmapped_device, for the Model column."""
    _record_unmapped(conn, "model_unmapped", "raw_model", raw_model)


# --- Cleaning report ------------------------------------------------------

@dataclass
class CleaningReport:
    source_file: str
    sheet_name: str = ""
    rows_read: int = 0
    rows_imported: int = 0
    rows_skipped_no_data: int = 0
    # [{"serial": "...", "asset_ids": [id, id, ...]}] - one entry per serial
    # that appeared more than once in this import, so each occurrence can be
    # opened directly for editing from the cleaning report.
    duplicate_serials: list[dict] = field(default_factory=list)
    unmapped_columns: list[str] = field(default_factory=list)
    unrecognized_devices: list[str] = field(default_factory=list)
    unrecognized_statuses: list[str] = field(default_factory=list)
    unrecognized_models: list[str] = field(default_factory=list)
    branch_hint: str = ""
    branch_matched: str = ""
    branch_no: str = ""
    batch_id: int = 0
    error: str = ""


def _ingest_asset_rows(
    conn,
    batch_id: int,
    rows,
    col_map: dict[str, int],
    branch_hint_default: str,
    source_label: str,
    report: CleaningReport,
    fixed_branch_no: str | None = None,
) -> None:
    """Normalize and insert a batch of equipment rows (shared by both the
    per-branch monthly report importer and the Total Asset history importer).

    When `fixed_branch_no` is given, every row is stamped with that branch_no
    instead of being resolved per row. A single monthly report file always
    covers one branch, and per-row "BRANCH/DEPT" text is often just a terse
    city nickname (e.g. "HANOI") that fuzzy-matches ambiguously against
    unrelated head-office departments containing the same city name - the
    file-level "Branch/TO/Center Name:" label (already resolved once by the
    caller) is the reliable signal. Per-row resolution is only used for the
    multi-branch Total Asset history import, where there is no such
    file-level label and each row genuinely can be a different branch.
    """
    seen_keys: dict[str, int] = {}  # asset_key -> first asset_item id seen
    duplicate_ids: dict[str, list[int]] = {}  # asset_key -> [asset_item ids], only for repeated serials
    serial_display: dict[str, str] = {}  # asset_key -> serial text as first seen (for display)
    branch_cache: dict = {}
    device_cache: dict = {}
    status_cache: dict = {}
    model_cache: dict = {}
    unrecognized_devices: set[str] = set()
    unrecognized_statuses: set[str] = set()
    unrecognized_models: set[str] = set()

    known_device_aliases = {
        row["alias"] for row in conn.execute("SELECT alias FROM device_aliases").fetchall()
    }
    known_standard_names = {
        row["name"] for row in conn.execute("SELECT name FROM device_standard_names").fetchall()
    }
    known_status_aliases = {
        row["alias"] for row in conn.execute("SELECT alias FROM status_aliases").fetchall()
    }
    known_standard_statuses = {
        row["name"] for row in conn.execute("SELECT name FROM status_standard_names").fetchall()
    }
    known_model_aliases = {
        row["alias"] for row in conn.execute("SELECT alias FROM model_aliases").fetchall()
    }
    known_standard_models = {
        row["name"] for row in conn.execute("SELECT name FROM model_standard_names").fetchall()
    }
    # Same idea as device-name mapping, but for people: an asset report's
    # free-text "FULL NAME" column is whatever the branch typed (spacing,
    # diacritics, capitalization all vary), while the IDFromAither user list
    # is the authoritative record for that same person, keyed by User No.
    # Standardize on the Aither spelling whenever the row's user ID resolves
    # to a known user; the report's own text is kept as full_name_raw.
    user_names_by_norm: dict[str, str] = {}
    for u in conn.execute("SELECT user_no, user_name FROM users").fetchall():
        _, u_norm = normalize_user_id(u["user_no"])
        if u_norm and u["user_name"]:
            user_names_by_norm[u_norm] = u["user_name"].strip().upper()

    def get(row, field_name):
        idx = col_map.get(field_name)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows:
        if row is None or all(c is None or _clean_str(c) == "" for c in row):
            continue
        report.rows_read += 1

        device_name_raw = _clean_str(get(row, "device_name"))
        serial_tag = _clean_serial(get(row, "serial_tag"))
        branch_dept_cell = _clean_str(get(row, "branch_dept")) or branch_hint_default

        if not device_name_raw and not serial_tag:
            report.rows_skipped_no_data += 1
            continue

        device_clean = device_name_raw.upper()
        if device_clean and device_clean not in known_device_aliases and device_clean not in known_standard_names:
            unrecognized_devices.add(device_clean)
            record_unmapped_device(conn, device_clean)

        device_name = normalize_device_name(conn, device_name_raw, cache=device_cache)

        status_raw = _clean_str(get(row, "status"))
        status_clean = status_raw.upper()
        if status_clean and status_clean not in known_status_aliases and status_clean not in known_standard_statuses:
            unrecognized_statuses.add(status_clean)
            record_unmapped_status(conn, status_clean)
        status = normalize_status(conn, status_raw, cache=status_cache)

        model_device_raw = _clean_str(get(row, "model_device"))
        model_clean = model_device_raw.upper()
        if model_clean and model_clean not in known_model_aliases and model_clean not in known_standard_models:
            unrecognized_models.add(model_clean)
            record_unmapped_model(conn, model_clean)
        model_device = normalize_model_device(conn, model_device_raw, cache=model_cache)

        user_id_raw, user_id_norm = normalize_user_id(get(row, "user_id"))
        asset_key = _asset_key(branch_dept_cell, device_name, model_device, serial_tag, user_id_norm)
        if fixed_branch_no is not None:
            branch_no = fixed_branch_no
        else:
            branch_no, _ = resolve_branch(conn, branch_dept_cell, cache=branch_cache)
            if not branch_no:
                record_unresolved_branch(conn, branch_dept_cell)

        full_name_raw = _clean_str(get(row, "full_name"))
        full_name = user_names_by_norm.get(user_id_norm, full_name_raw.upper())

        cursor = conn.execute(
            """
            INSERT INTO asset_items (
                batch_id, asset_key, branch_dept, branch_no, device_name, device_name_raw,
                user_id_raw, user_id_norm, full_name, full_name_raw, model_device, model_device_raw,
                serial_tag, status, status_raw, remark, position, handover_date, ip, extra_json, source_file
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                asset_key,
                branch_dept_cell,
                branch_no,
                device_name,
                device_name_raw,
                user_id_raw,
                user_id_norm,
                full_name,
                full_name_raw,
                model_device,
                model_device_raw,
                serial_tag,
                status,
                status_raw,
                _clean_str(get(row, "remark")),
                _clean_str(get(row, "position")),
                _parse_date(get(row, "handover_date")),
                _clean_str(get(row, "ip")),
                None,
                source_label,
            ),
        )
        new_id = cursor.lastrowid
        report.rows_imported += 1

        if serial_tag:
            if asset_key in seen_keys:
                ids = duplicate_ids.setdefault(asset_key, [seen_keys[asset_key]])
                ids.append(new_id)
            else:
                seen_keys[asset_key] = new_id
                serial_display[asset_key] = serial_tag

    report.unrecognized_devices = sorted(unrecognized_devices)
    report.unrecognized_statuses = sorted(unrecognized_statuses)
    report.unrecognized_models = sorted(unrecognized_models)
    report.duplicate_serials = []
    for key, ids in sorted(duplicate_ids.items()):
        placeholders = ",".join("?" * len(ids))
        # Full row details (not just the serial + ids) so the Cleaning
        # Report can show exactly what's duplicated - same device handed to
        # two different people, a copy-pasted row, etc. - instead of making
        # the user open each Edit link just to see what they're comparing.
        detail_rows = conn.execute(
            f"SELECT * FROM asset_items WHERE id IN ({placeholders})", ids
        ).fetchall()
        report.duplicate_serials.append(
            {"serial": serial_display.get(key, key), "asset_ids": ids, "rows": detail_rows}
        )


def peek_asset_report_branch(conn, path: str) -> dict:
    """Lightweight pre-import check: which branch a report file would
    resolve to, without creating a batch or inserting any rows. Used by
    routes/import_data.py to warn the user if that branch already has an
    import for the period they're about to use, before actually importing -
    re-opens/re-parses the file rather than sharing code with
    import_asset_report, so the real ingest path can't be affected by
    changes made for this peek."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - just means the check is skipped, real import will report it
        return {"branch_no": "", "branch_matched": "", "branch_hint": ""}

    match = detect_equipment_sheet(wb)
    if not match:
        wb.close()
        return {"branch_no": "", "branch_matched": "", "branch_hint": ""}

    branch_hint = match.branch_hint
    if not branch_hint and "branch_dept" in match.col_map:
        ws = wb[match.sheet_name]
        idx = match.col_map["branch_dept"]
        for row in ws.iter_rows(min_row=match.header_row_idx + 2, values_only=True):
            if row is None or idx >= len(row):
                continue
            value = _clean_str(row[idx])
            if value:
                branch_hint = value
                break

    branch_no, branch_matched = resolve_branch(conn, branch_hint)
    wb.close()
    return {"branch_no": branch_no, "branch_matched": branch_matched, "branch_hint": branch_hint}


def import_asset_report(path: str, source_label: str | None = None, period: str | None = None) -> CleaningReport:
    report = CleaningReport(source_file=source_label or path)
    conn = get_connection()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced to the user as-is
        report.error = f"Could not open file: {exc}"
        log_import(conn, "asset_report", report.source_file, period=period or "", result=report.error)
        conn.commit()
        conn.close()
        return report

    match = detect_equipment_sheet(wb)
    if not match:
        wb.close()
        report.error = "No equipment list sheet found (no header row matched DEVICE NAME + SERIAL columns)."
        log_import(conn, "asset_report", report.source_file, period=period or "", result=report.error)
        conn.commit()
        conn.close()
        return report

    report.sheet_name = match.sheet_name
    report.branch_hint = match.branch_hint

    ws = wb[match.sheet_name]

    if not report.branch_hint and "branch_dept" in match.col_map:
        # No separate "Branch/TO/Center Name:" label row was found - some
        # files (e.g. ones with a plain "BRANCH" column instead) only carry
        # the branch name once per row. Fall back to the first non-empty
        # value in that column so the file still resolves to a branch
        # instead of every row silently landing in "unresolved".
        idx = match.col_map["branch_dept"]
        for row in ws.iter_rows(min_row=match.header_row_idx + 2, values_only=True):
            if row is None or idx >= len(row):
                continue
            value = _clean_str(row[idx])
            if value:
                report.branch_hint = value
                break

    branch_no, branch_matched = resolve_branch(conn, report.branch_hint)
    report.branch_matched = branch_matched
    report.branch_no = branch_no
    if not branch_no:
        record_unresolved_branch(conn, report.branch_hint)

    header_row_values = next(
        ws.iter_rows(min_row=match.header_row_idx + 1, max_row=match.header_row_idx + 1, values_only=True)
    )
    known_col_idxs = set(match.col_map.values())
    for idx, cell in enumerate(header_row_values):
        text = _clean_str(cell)
        if text and idx not in known_col_idxs:
            report.unmapped_columns.append(text)

    now = _now_iso()
    resolved_period = period.strip() if period and period.strip() else now[:7]
    batch_id = conn.execute(
        "INSERT INTO import_batches (imported_at, kind, source_files_json, label, period) VALUES (?, 'asset_report', ?, ?, ?)",
        (now, json.dumps([report.source_file]), report.branch_hint or match.sheet_name, resolved_period),
    ).lastrowid
    report.batch_id = batch_id

    data_rows = ws.iter_rows(min_row=match.header_row_idx + 2, values_only=True)
    _ingest_asset_rows(
        conn, batch_id, data_rows, match.col_map, match.branch_hint, report.source_file, report,
        fixed_branch_no=branch_no,
    )

    log_import(conn, "asset_report", report.source_file, rows_processed=report.rows_imported, period=resolved_period)
    conn.commit()
    conn.close()
    wb.close()

    return report


def _month_sort_key(month_text: str, year) -> tuple:
    m = re.search(r"\d+", _clean_str(month_text))
    month_num = int(m.group()) if m else 0
    try:
        year_num = int(year)
    except (TypeError, ValueError):
        year_num = 0
    return (year_num, month_num)


def import_total_asset_history(path: str) -> list[CleaningReport]:
    """Load the aggregated 'Total Asset' workbook, which packs several months of
    every branch's equipment list into one flat sheet (with Month/Years columns
    instead of one file per month). Each distinct (Month, Years) becomes its own
    historical import_batch, inserted oldest-first, using the same row
    normalization as a regular monthly asset report."""
    wb = openpyxl.load_workbook(path, data_only=True)
    match = detect_equipment_sheet(wb)
    if not match:
        wb.close()
        error = "No equipment list sheet found in Total Asset file."
        conn = get_connection()
        log_import(conn, "total_asset_baseline", path, result=error)
        conn.commit()
        conn.close()
        return [CleaningReport(source_file=path, error=error)]

    ws = wb[match.sheet_name]
    header_row_values = next(
        ws.iter_rows(min_row=match.header_row_idx + 1, max_row=match.header_row_idx + 1, values_only=True)
    )
    month_idx = year_idx = None
    for idx, cell in enumerate(header_row_values):
        text = _normalize_header_cell(cell)
        if text == "MONTH":
            month_idx = idx
        elif text in ("YEARS", "YEAR"):
            year_idx = idx

    groups: dict[tuple, list] = {}
    for row in ws.iter_rows(min_row=match.header_row_idx + 2, values_only=True):
        if row is None or all(c is None or _clean_str(c) == "" for c in row):
            continue
        month_text = row[month_idx] if month_idx is not None and month_idx < len(row) else None
        year_val = row[year_idx] if year_idx is not None and year_idx < len(row) else None
        if not _clean_str(month_text) or not _clean_str(year_val):
            continue
        groups.setdefault((_clean_str(month_text), _clean_str(year_val)), []).append(row)

    conn = get_connection()
    reports: list[CleaningReport] = []
    now = _now_iso()
    for (month_text, year_val), rows in sorted(groups.items(), key=lambda kv: _month_sort_key(*kv[0])):
        label = f"{month_text} {year_val} (Total Asset baseline)"
        report = CleaningReport(source_file=f"{path} :: {label}", sheet_name=match.sheet_name)
        year_num, month_num = _month_sort_key(month_text, year_val)
        period = f"{year_num:04d}-{month_num:02d}"
        batch_id = conn.execute(
            "INSERT INTO import_batches (imported_at, kind, source_files_json, label, period) VALUES (?, 'asset_report', ?, ?, ?)",
            (now, json.dumps([path]), label, period),
        ).lastrowid
        report.batch_id = batch_id
        _ingest_asset_rows(conn, batch_id, rows, match.col_map, "", report.source_file, report)
        log_import(conn, "total_asset_baseline", report.source_file, rows_processed=report.rows_imported, period=period)
        reports.append(report)

    conn.commit()
    conn.close()
    wb.close()
    return reports


# --- ID master files (branches / users) -----------------------------------

def import_branch_file(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {_normalize_header_cell(h): idx for idx, h in enumerate(header) if h}

    required = ["BRANCH NO", "LOCAL BRANCH NAME", "ENG. BRANCH NAME"]
    if not all(r in col for r in required):
        wb.close()
        error = f"Unrecognized branch file layout: {header}"
        conn = get_connection()
        log_import(conn, "branch_codes", path, result=error)
        conn.commit()
        conn.close()
        return {"error": error}

    conn = get_connection()
    now = _now_iso()
    count = 0
    for row in rows:
        branch_no = _clean_str(row[col["BRANCH NO"]])
        if not branch_no:
            continue
        local_name = _clean_str(row[col["LOCAL BRANCH NAME"]])
        # Every branch's eng_name repeats "SHINHAN BANK VIETNAM(E)" as a
        # prefix - redundant everywhere it's displayed since the whole app
        # is already scoped to this one bank, so it's stripped at the
        # source rather than patched at every display site.
        eng_name = strip_bank_prefix(_clean_str(row[col["ENG. BRANCH NAME"]]))
        conn.execute(
            """
            INSERT INTO branches (branch_no, local_name, eng_name, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(branch_no) DO UPDATE SET
                local_name = excluded.local_name,
                eng_name = excluded.eng_name,
                updated_at = excluded.updated_at
            """,
            (branch_no, local_name, eng_name, now),
        )
        count += 1
    log_import(conn, "branch_codes", path, rows_processed=count)
    conn.commit()
    conn.close()
    wb.close()
    return {"rows": count}


def import_user_file(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {_normalize_header_cell(h): idx for idx, h in enumerate(header) if h}

    required = ["BRANCH ID", "USER NO", "USER NAME"]
    if not all(r in col for r in required):
        wb.close()
        error = f"Unrecognized user file layout: {header}"
        conn = get_connection()
        log_import(conn, "user_ids", path, result=error)
        conn.commit()
        conn.close()
        return {"error": error}

    conn = get_connection()
    now = _now_iso()
    count = 0
    for row in rows:
        user_no = _clean_str(row[col["USER NO"]])
        if not user_no:
            continue
        branch_no = _clean_str(row[col["BRANCH ID"]])
        user_name = _clean_str(row[col["USER NAME"]])
        eng_name = _clean_str(row[col["ENG. BANKER NAME"]]) if "ENG. BANKER NAME" in col else ""
        banker_key = _clean_str(row[col["BANKER KEY NUMBER"]]) if "BANKER KEY NUMBER" in col else ""
        status = _clean_str(row[col["STATUS"]]) if "STATUS" in col else ""
        _, user_no_norm = normalize_user_id(user_no)
        conn.execute(
            """
            INSERT INTO users (user_no, user_no_norm, branch_no, user_name, eng_name, banker_key, status, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(user_no) DO UPDATE SET
                user_no_norm = excluded.user_no_norm,
                branch_no = excluded.branch_no,
                user_name = excluded.user_name,
                eng_name = excluded.eng_name,
                banker_key = excluded.banker_key,
                status = excluded.status,
                updated_at = excluded.updated_at
            """,
            (user_no, user_no_norm, branch_no, user_name, eng_name, banker_key, status, now),
        )
        count += 1
    log_import(conn, "user_ids", path, rows_processed=count)
    conn.commit()
    conn.close()
    wb.close()
    return {"rows": count}


def import_id_file(path: str) -> dict:
    """Detect whether a file from IDFromAither/ is a branch list or a user list."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    keys = {_normalize_header_cell(h) for h in header if h}
    wb.close()

    if "LOCAL BRANCH NAME" in keys:
        result = import_branch_file(path)
        result["file_type"] = "branches"
        return result
    if "USER NO" in keys:
        result = import_user_file(path)
        result["file_type"] = "users"
        return result
    return {"error": f"Unrecognized ID file layout (header: {sorted(keys)})", "file_type": "unknown"}


def find_user(conn, query: str):
    """Resolve free-text user-search input (typed exactly, or missing a
    leading zero, or entered as a bare int) to a users-table row. Shared by
    Lookup and User History so both pages match IDs the same way.

    The exact/normalized-verbatim attempts below cover the common case with
    a plain PK lookup; the fallback matches via the indexed `user_no_norm`
    column (populated at import time and backfilled for any pre-existing
    database - see db.py) instead of pulling and normalizing every row in
    Python, which used to mean a full ~8000-row table scan on every search
    that didn't hit an exact format match."""
    query = (query or "").strip()
    if not query:
        return None
    _, norm = normalize_user_id(query)
    row = conn.execute("SELECT * FROM users WHERE user_no = ?", (query,)).fetchone()
    if not row and norm:
        row = conn.execute("SELECT * FROM users WHERE user_no = ?", (norm,)).fetchone()
    if not row and norm:
        row = conn.execute("SELECT * FROM users WHERE user_no_norm = ? LIMIT 1", (norm,)).fetchone()
    return row
