import io
import math

import openpyxl
from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from openpyxl.styles import Font

from ..db import get_connection
from ..queries import (
    UNRESOLVED_BRANCH_FILTER,
    find_current_duplicate_serials,
    get_branch,
    get_branches_with_current_assets,
    has_unresolved_current_assets,
    search_assets,
)

bp = Blueprint("asset_edit", __name__, url_prefix="/assets")

PAGE_SIZE = 100

# Free-text fields a user can correct by hand (e.g. to resolve a duplicate
# serial flagged by the cleaning report). branch_no/batch_id/asset_key are
# deliberately not editable here - they're derived by the importer, not
# something to hand-edit.
EDITABLE_FIELDS = [
    "device_name",
    "model_device",
    "serial_tag",
    "status",
    "remark",
    "position",
    "branch_dept",
    "full_name",
    "user_id_raw",
    "handover_date",
    "ip",
]

# Every editable field is uppercased except handover_date, which is a date,
# not free text - matches the all-caps convention used everywhere else data
# ends up on a printed form or a report (see lookup.py's UPPERCASE_FIELDS).
UPPERCASE_FIELDS = set(EDITABLE_FIELDS) - {"handover_date"}


@bp.route("/")
def index():
    filters = {
        "branch_no": [v for v in request.args.getlist("branch_no") if v],
        "device_name": [v for v in request.args.getlist("device_name") if v],
        "status": [v for v in request.args.getlist("status") if v],
        "q": request.args.get("q", "").strip(),
    }
    try:
        page = max(int(request.args.get("page", 1)), 1)
    except ValueError:
        page = 1

    conn = get_connection()
    try:
        rows, total = search_assets(conn, filters, page=page, per_page=PAGE_SIZE)
        branches = get_branches_with_current_assets(conn)
        show_unresolved_option = has_unresolved_current_assets(conn)
        device_names = [
            r["device_name"]
            for r in conn.execute(
                "SELECT DISTINCT device_name FROM asset_items WHERE device_name != '' ORDER BY device_name"
            ).fetchall()
        ]
        # Union of the standard status list and whatever status values are
        # actually in use right now, so the filter covers legacy/non-standard
        # values too instead of hiding them the moment a standard list exists.
        current_statuses = {
            r["status"]
            for r in conn.execute(
                "SELECT DISTINCT status FROM asset_items WHERE status != ''"
            ).fetchall()
        }
        standard_statuses = {
            r["name"] for r in conn.execute("SELECT name FROM status_standard_names").fetchall()
        }
        status_options = sorted(current_statuses | standard_statuses)
        # Only worth showing the "filtered to this one branch" convenience
        # message/link (see template) when exactly one branch is selected -
        # with several selected at once there's no single branch to link to.
        selected_branch = None
        if len(filters["branch_no"]) == 1:
            only_branch = filters["branch_no"][0]
            if only_branch == UNRESOLVED_BRANCH_FILTER:
                selected_branch = {"branch_no": "", "eng_name": "Unresolved / unmatched branch"}
            else:
                selected_branch = get_branch(conn, only_branch)
    finally:
        conn.close()

    total_pages = max(math.ceil(total / PAGE_SIZE), 1)

    return render_template(
        "asset_management.html",
        active_page="assets",
        rows=rows,
        total=total,
        page=page,
        total_pages=total_pages,
        filters=filters,
        branches=branches,
        show_unresolved_option=show_unresolved_option,
        unresolved_value=UNRESOLVED_BRANCH_FILTER,
        device_names=device_names,
        status_options=status_options,
        selected_branch=selected_branch,
    )


@bp.route("/duplicates")
def duplicates():
    """On-demand, system-wide duplicate-serial check across every branch's
    current assets - broader than the automatic check shown right after an
    import, which only looks within that one file's own rows."""
    conn = get_connection()
    try:
        dupes = find_current_duplicate_serials(conn)
    finally:
        conn.close()
    return render_template("duplicate_check.html", active_page="assets", dupes=dupes)


def build_duplicates_workbook(dupes: list) -> openpyxl.Workbook:
    """Shared by both places duplicates get exported: the system-wide
    Duplicate Check page (dupes from find_current_duplicate_serials) and the
    per-import Cleaning Report (dupes from find_duplicate_serials_in_batch,
    called from import_data.export_duplicates) - same [{"serial", "rows"}]
    shape either way, just missing the joined branch_eng_name column in the
    per-import case since that query has no branches JOIN."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicates"
    ws.append(["Serial", "Branch", "Device", "Model", "Full Name", "User ID", "Status", "Asset ID"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for d in dupes:
        for row in d["rows"]:
            branch = row["branch_eng_name"] if "branch_eng_name" in row.keys() else ""
            ws.append(
                [
                    d["serial"], branch or row["branch_dept"], row["device_name"], row["model_device"],
                    row["full_name"], row["user_id_raw"], row["status"], row["id"],
                ]
            )

    widths = [16, 26, 14, 18, 22, 14, 12, 10]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    return wb


@bp.route("/duplicates/export")
def export_duplicates():
    conn = get_connection()
    try:
        dupes = find_current_duplicate_serials(conn)
    finally:
        conn.close()
    if not dupes:
        flash("No duplicate serials found among current assets.", "error")
        return redirect(url_for("asset_edit.duplicates"))
    wb = build_duplicates_workbook(dupes)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="duplicate_assets.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/bulk-delete", methods=["POST"])
def bulk_delete():
    ids = [int(x) for x in request.form.getlist("asset_ids") if x.isdigit()]
    if not ids:
        flash("No assets selected.", "error")
    else:
        conn = get_connection()
        try:
            placeholders = ",".join("?" * len(ids))
            conn.execute(f"DELETE FROM asset_items WHERE id IN ({placeholders})", ids)
            conn.commit()
            flash(f"Deleted {len(ids)} asset(s).", "success")
        finally:
            conn.close()
    next_url = request.form.get("next") or url_for("dashboard.index")
    return redirect(next_url)


@bp.route("/<int:asset_id>/delete", methods=["POST"])
def delete(asset_id):
    conn = get_connection()
    try:
        row = conn.execute("SELECT id FROM asset_items WHERE id = ?", (asset_id,)).fetchone()
        if not row:
            flash("Asset not found.", "error")
        else:
            conn.execute("DELETE FROM asset_items WHERE id = ?", (asset_id,))
            conn.commit()
            flash(f"Asset #{asset_id} deleted.", "success")
    finally:
        conn.close()
    next_url = request.form.get("next") or url_for("dashboard.index")
    return redirect(next_url)


@bp.route("/<int:asset_id>/edit", methods=["GET", "POST"])
def edit(asset_id):
    conn = get_connection()
    try:
        if request.method == "POST":
            asset = conn.execute("SELECT id FROM asset_items WHERE id = ?", (asset_id,)).fetchone()
            if not asset:
                flash("Asset not found.", "error")
                return redirect(url_for("dashboard.index"))

            values = {}
            for field_name in EDITABLE_FIELDS:
                value = request.form.get(field_name, "").strip()
                values[field_name] = value.upper() if field_name in UPPERCASE_FIELDS else value

            set_clause = ", ".join(f"{f} = ?" for f in EDITABLE_FIELDS)
            conn.execute(
                f"UPDATE asset_items SET {set_clause} WHERE id = ?",
                [*values.values(), asset_id],
            )
            conn.commit()
            flash(f"Asset #{asset_id} updated.", "success")
            next_url = request.form.get("next") or url_for("dashboard.index")
            return redirect(next_url)

        asset = conn.execute("SELECT * FROM asset_items WHERE id = ?", (asset_id,)).fetchone()
        if not asset:
            flash("Asset not found.", "error")
            return redirect(url_for("dashboard.index"))
        batch = conn.execute(
            "SELECT * FROM import_batches WHERE id = ?", (asset["batch_id"],)
        ).fetchone()
        standard_names = [
            r["name"] for r in conn.execute("SELECT name FROM device_standard_names ORDER BY name").fetchall()
        ]
        standard_statuses = [
            r["name"] for r in conn.execute("SELECT name FROM status_standard_names ORDER BY name").fetchall()
        ]
    finally:
        conn.close()

    next_url = request.args.get("next", "")
    return render_template(
        "asset_edit.html",
        active_page="assets",
        asset=asset,
        batch=batch,
        standard_names=standard_names,
        standard_statuses=standard_statuses,
        next_url=next_url,
    )
