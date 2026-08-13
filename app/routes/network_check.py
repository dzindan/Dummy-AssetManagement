"""Live network verification: pings every IP recorded against a branch's
current assets, pulls each live machine's PC serial / monitor serial /
logged-on user over the network (reusing app/scanner.py, lifted as-is from
the standalone IP Scanner tool - ping + `quser` + PowerShell Get-WmiObject,
no extra dependencies), and compares that against what was imported - so a
mismatch (wrong serial recorded, different person logged in than the asset
is assigned to) surfaces without a manual walkthrough.

This only works for machines reachable from wherever this app is running
(same LAN/VPN, WMI/RPC not blocked by firewall, an account with admin rights
on the target) - see the note on the Network Check page itself. A branch
being unreachable just shows every IP as "No response", same as the
original standalone scanner.
"""

from __future__ import annotations

import io
import re
import threading
import uuid

import openpyxl
from flask import Blueprint, jsonify, render_template, request, send_file
from openpyxl.styles import Font

from ..db import get_connection
from ..paths import safe_filename
from ..queries import get_branch, get_branches_with_current_assets, get_current_assets
from ..scanner import DEFAULT_CONCURRENCY, run_scan

bp = Blueprint("network_check", __name__, url_prefix="/network-check")

_SCANS: dict[str, dict] = {}
_LOCK = threading.Lock()

PC_DEVICE_NAMES = {"PC", "NOTEBOOK", "SERVER PC"}
MONITOR_DEVICE_NAMES = {"LCD"}


def _normalize_serial(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def _live_pc_serial(hardware: dict | None) -> str:
    if not hardware:
        return ""
    system = hardware.get("system") or {}
    if not system.get("ok"):
        return ""
    items = system.get("items") or []
    return (items[0].get("Serial") or "").strip() if items else ""


def _live_monitor_serials(hardware: dict | None) -> list[str]:
    if not hardware:
        return []
    monitors = hardware.get("monitors") or {}
    if not monitors.get("ok"):
        return []
    return [(m.get("SerialNumber") or "").strip() for m in (monitors.get("items") or [])]


def _live_logon_users(sessions: dict | None) -> list[str]:
    if not sessions or not sessions.get("ok"):
        return []
    users = []
    for s in sessions.get("sessions") or []:
        username = s.get("username") or ""
        if "\\" in username:
            username = username.split("\\", 1)[1]
        users.append(username)
    return users


def _live_mac(hardware: dict | None) -> str:
    if not hardware:
        return ""
    mac_address = hardware.get("mac_address") or {}
    if not mac_address.get("ok"):
        return ""
    return mac_address.get("mac") or ""


UPDATABLE_FIELDS = {
    # field key -> (asset_items column to write)
    "pc_serial": "serial_tag",
    "monitor_serial": "serial_tag",
    "user": "user_id_raw",
}


def compare_result(live: dict, imported_rows: list) -> dict:
    """Cross-check one scanned IP's live hardware/session info against every
    imported asset row recorded at that same IP (there can be several - a
    PC, its monitor, sharing one person's IP). Also carries the asset_items
    id(s) each field would need to write to if the user clicks "Update" on a
    mismatch - see apply_updates() below."""
    imported_pc = next((r for r in imported_rows if r["device_name"] in PC_DEVICE_NAMES and r["serial_tag"]), None)
    imported_monitor = next(
        (r for r in imported_rows if r["device_name"] in MONITOR_DEVICE_NAMES and r["serial_tag"]), None
    )
    imported_user_row = next((r for r in imported_rows if r["user_id_norm"] or r["full_name"]), None)

    live_pc_serial = _live_pc_serial(live.get("hardware"))
    live_monitor_serials = _live_monitor_serials(live.get("hardware"))
    live_users = _live_logon_users(live.get("sessions"))

    # Each *_match stays None ("N/A") whenever there's nothing live to
    # compare - either nothing was imported for that field, or the live read
    # itself failed/came back empty (offline machine, WMI blocked by
    # firewall, no monitor detected...). Only set True/False once there's an
    # actual live value to weigh against the imported one - otherwise "we
    # couldn't read it" would show as a false MISMATCH with no live value to
    # offer an Update for for (see UPDATABLE_FIELDS in network_check.js,
    # which already requires a live value before showing the button - this
    # keeps the badge honest about the same distinction).
    pc_match = None
    if imported_pc and live_pc_serial:
        pc_match = _normalize_serial(live_pc_serial) == _normalize_serial(imported_pc["serial_tag"])

    monitor_match = None
    if imported_monitor and live_monitor_serials:
        target = _normalize_serial(imported_monitor["serial_tag"])
        monitor_match = any(_normalize_serial(s) == target for s in live_monitor_serials)

    user_match = None
    if imported_user_row and (imported_user_row["user_id_norm"] or imported_user_row["user_id_raw"]) and live_users:
        target_id = (imported_user_row["user_id_norm"] or imported_user_row["user_id_raw"] or "").upper()
        user_match = any(u.upper() == target_id for u in live_users)

    return {
        "imported_pc_serial": imported_pc["serial_tag"] if imported_pc else "",
        "live_pc_serial": live_pc_serial,
        "pc_match": pc_match,
        "pc_asset_ids": [imported_pc["id"]] if imported_pc else [],
        "imported_monitor_serial": imported_monitor["serial_tag"] if imported_monitor else "",
        "live_monitor_serials": live_monitor_serials,
        "monitor_match": monitor_match,
        "monitor_asset_ids": [imported_monitor["id"]] if imported_monitor else [],
        "imported_user": imported_user_row["full_name"] if imported_user_row else "",
        "imported_user_id": (imported_user_row["user_id_norm"] or imported_user_row["user_id_raw"]) if imported_user_row else "",
        "live_users": live_users,
        "user_match": user_match,
        # A mismatched user is corrected on every device at this IP, not just
        # one row - see the module docstring: a PC and its monitor share one
        # person's IP, so they should share the same recorded user too.
        "user_asset_ids": [r["id"] for r in imported_rows],
        # No MAC address is imported anywhere in the asset data, so this is
        # informational only (like hostname) - not something to MATCH/MISMATCH.
        "live_mac": _live_mac(live.get("hardware")),
    }


@bp.route("/")
def index():
    conn = get_connection()
    try:
        branches = get_branches_with_current_assets(conn)
    finally:
        conn.close()
    return render_template("network_check.html", active_page="network_check", branches=branches)


@bp.route("/scan", methods=["POST"])
def start_scan():
    body = request.get_json(silent=True) or {}
    branch_no = (body.get("branch_no") or "").strip()
    if not branch_no:
        return jsonify({"error": "Select a branch first."}), 400

    conn = get_connection()
    try:
        rows = get_current_assets(conn, branch_no=branch_no)
        branch_row = get_branch(conn, branch_no)
    finally:
        conn.close()

    # Plain dicts (not sqlite3.Row) so an applied update can mutate a row
    # in place and have the next poll of this same scan reflect it, without
    # a full re-scan.
    snapshot: dict[str, list] = {}
    for row in rows:
        ip = (row["ip"] or "").strip()
        if ip:
            snapshot.setdefault(ip, []).append(dict(row))

    targets = sorted(snapshot.keys())
    if not targets:
        return jsonify({"error": "This branch has no current assets with an IP address recorded."}), 400

    scan_id = uuid.uuid4().hex
    with _LOCK:
        _SCANS[scan_id] = {
            "total": len(targets),
            "done": 0,
            "results": [],
            "order": targets,
            "snapshot": snapshot,
            "branch_no": branch_no,
            "branch_label": branch_row["eng_name"] if branch_row else branch_no,
            "stop_requested": False,
            "worker_done": False,
        }

    def should_stop() -> bool:
        with _LOCK:
            return _SCANS[scan_id]["stop_requested"]

    def worker():
        def on_result(result: dict):
            with _LOCK:
                state = _SCANS[scan_id]
                state["results"].append(result)
                state["done"] += 1

        run_scan(targets, on_result, max_workers=DEFAULT_CONCURRENCY, include_hardware=True, should_stop=should_stop)
        with _LOCK:
            _SCANS[scan_id]["worker_done"] = True

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"scan_id": scan_id, "total": len(targets), "branch_label": _SCANS[scan_id]["branch_label"]})


@bp.route("/scan/<scan_id>/stop", methods=["POST"])
def stop_scan(scan_id: str):
    with _LOCK:
        state = _SCANS.get(scan_id)
        if state is None:
            return jsonify({"error": "Scan session not found."}), 404
        state["stop_requested"] = True
    return jsonify({"ok": True})


@bp.route("/scan/<scan_id>/apply", methods=["POST"])
def apply_updates(scan_id: str):
    """Write one or more mismatched fields' live-scanned value into the
    database, e.g. after the user clicks "Update" on a single mismatched
    cell, or "Update All" for every mismatch currently shown. Every actual
    change is logged to network_check_log for an audit trail; a value that
    already matches the current DB row is skipped (no-op, not logged)."""
    with _LOCK:
        state = _SCANS.get(scan_id)
        if state is None:
            return jsonify({"error": "Scan session not found."}), 404
        snapshot = state["snapshot"]
        branch_no = state["branch_no"]

    body = request.get_json(silent=True) or {}
    updates = body.get("updates") or []

    applied = []
    conn = get_connection()
    try:
        for u in updates:
            ip = (u.get("ip") or "").strip()
            field = u.get("field")
            new_value = (u.get("value") or "").strip().upper()
            asset_ids = u.get("asset_ids") or []
            column = UPDATABLE_FIELDS.get(field)
            if not (ip and column and asset_ids and new_value):
                continue

            for asset_id in asset_ids:
                row = conn.execute(f"SELECT {column} FROM asset_items WHERE id = ?", (asset_id,)).fetchone()
                if row is None or row[column] == new_value:
                    continue
                old_value = row[column]
                conn.execute(f"UPDATE asset_items SET {column} = ? WHERE id = ?", (new_value, asset_id))
                conn.execute(
                    "INSERT INTO network_check_log "
                    "(applied_at, branch_no, ip, asset_id, field, old_value, new_value) "
                    "VALUES (datetime('now'), ?, ?, ?, ?, ?, ?)",
                    (branch_no, ip, asset_id, field, old_value, new_value),
                )
                applied.append({"ip": ip, "field": field, "asset_id": asset_id, "new_value": new_value})

                with _LOCK:
                    for r in snapshot.get(ip, []):
                        if r["id"] == asset_id:
                            r[column] = new_value
        conn.commit()
    finally:
        conn.close()

    return jsonify({"applied": applied})


def _augmented_results(state: dict) -> list[dict]:
    order = {ip: i for i, ip in enumerate(state["order"])}
    results = sorted(state["results"], key=lambda r: order.get(r["ip"], 0))
    augmented = []
    for r in results:
        imported_rows = state["snapshot"].get(r["ip"], [])
        augmented.append({**r, "compare": compare_result(r, imported_rows)})
    return augmented


@bp.route("/scan/<scan_id>")
def scan_status(scan_id: str):
    with _LOCK:
        state = _SCANS.get(scan_id)
        if state is None:
            return jsonify({"error": "Scan session not found."}), 404
        results = _augmented_results(state)
        return jsonify(
            {
                "total": state["total"],
                "done": state["done"],
                "finished": state["worker_done"],
                "stopped": state["stop_requested"],
                "branch_label": state["branch_label"],
                "results": results,
            }
        )


def _match_text(value: bool | None) -> str:
    if value is None:
        return "N/A"
    return "MATCH" if value else "MISMATCH"


@bp.route("/scan/<scan_id>/export.xlsx")
def export_xlsx(scan_id: str):
    with _LOCK:
        state = _SCANS.get(scan_id)
        if state is None:
            return jsonify({"error": "Scan session not found."}), 404
        results = _augmented_results(state)
        branch_label = state["branch_label"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Network Check"
    headers = [
        "IP", "Alive", "Hostname", "MAC Address",
        "PC Serial (Live)", "PC Serial (Imported)", "PC Match",
        "Monitor Serial (Live)", "Monitor Serial (Imported)", "Monitor Match",
        "Logged-on User (Live)", "User (Imported)", "User Match",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in results:
        c = r["compare"]
        ws.append(
            [
                r["ip"],
                "Yes" if r["alive"] else "No",
                r.get("hostname") or "",
                c["live_mac"],
                c["live_pc_serial"],
                c["imported_pc_serial"],
                _match_text(c["pc_match"]),
                ", ".join(c["live_monitor_serials"]),
                c["imported_monitor_serial"],
                _match_text(c["monitor_match"]),
                ", ".join(c["live_users"]),
                c["imported_user"],
                _match_text(c["user_match"]),
            ]
        )

    widths = [16, 8, 20, 18, 22, 22, 12, 22, 22, 14, 22, 22, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_branch = safe_filename(branch_label, fallback="branch")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"network_check_{safe_branch}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/log")
def log():
    conn = get_connection()
    try:
        rows = conn.execute(
            """
            SELECT ncl.*, b.eng_name AS branch_eng_name
            FROM network_check_log ncl
            LEFT JOIN branches b ON b.branch_no = ncl.branch_no
            ORDER BY ncl.id DESC LIMIT 500
            """
        ).fetchall()
    finally:
        conn.close()
    return render_template("network_check_log.html", active_page="network_check", rows=rows)
