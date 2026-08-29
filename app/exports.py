"""Shared Excel-export plumbing - the only module that touches openpyxl
directly. Every export/template-download route builds its data the normal
way (a query, a computed table, ...) and then calls into here to turn it
into a downloadable .xlsx, instead of constructing a Workbook inline."""

from __future__ import annotations

import datetime as dt
import io
from typing import Any, Callable, Iterable, Sequence

import openpyxl
from flask import Response, send_file
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from .diffing import COMPARE_FIELDS, FIELD_LABELS
from .text_utils import usage_duration_years

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def dated_download_name(base: str, *, with_time: bool = False) -> str:
    """`base` (no extension) -> "{base}_{today}.xlsx" so a downloaded report
    carries the date it was generated - repeated exports of the same report
    type don't silently overwrite each other in a Downloads folder, and a
    file is identifiable on its own once it's sitting next to others.
    `with_time=True` also appends a time-of-day, for reports (live network/
    CUCM scans) that can reasonably be re-run and re-exported several times
    within the same day."""
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S") if with_time else dt.date.today().isoformat()
    return f"{base}_{stamp}.xlsx"


def match_text(value: bool | None) -> str:
    """Renders one of the tri-state MATCH/MISMATCH/N/A comparison flags
    live-scan cross-checks produce (Network Check's pc_match/monitor_match/
    user_match, CUCM Phone Scan's model_match/serial_match/user_match - see
    each module's compare_result()/_with_import_crosscheck()) as export
    text. None means "nothing live to compare against" (see those
    functions' own docstrings on when that happens), not a mismatch."""
    if value is None:
        return "N/A"
    return "MATCH" if value else "MISMATCH"


# A column is (header_label, key_or_getter). A string key reads row[key]
# (works for both sqlite3.Row and dict); a callable is called as getter(row)
# for computed/joined/formatted values.
ColumnSpec = tuple[str, "str | Callable[[Any], Any]"]


def _cell_value(row: Any, key_or_getter) -> Any:
    if callable(key_or_getter):
        return key_or_getter(row)
    return row[key_or_getter]


def style_header_row(ws: Worksheet, widths: Sequence[int] | None = None) -> None:
    """Bold row 1, plus a column width - explicit if given, else derived
    from the header label's own length. Applied everywhere so every
    exported sheet looks consistent (previously only 2 of 4 export
    functions styled their header at all)."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, cell in enumerate(ws[1], start=1):
        width = widths[i - 1] if widths and i - 1 < len(widths) else max(10, min(30, len(str(cell.value or "")) + 4))
        ws.column_dimensions[cell.column_letter].width = width


def write_sheet(
    ws: Worksheet, columns: Sequence[ColumnSpec], rows: Iterable, widths: Sequence[int] | None = None
) -> None:
    """Append a header row (column labels) then one row per item, then style
    the header - the one place row-shaped data becomes Excel cells."""
    ws.append([c[0] for c in columns])
    for row in rows:
        ws.append([_cell_value(row, c[1]) for c in columns])
    style_header_row(ws, widths)


def build_workbook(
    sheet_title: str, columns: Sequence[ColumnSpec], rows: Iterable, widths: Sequence[int] | None = None
) -> openpyxl.Workbook:
    """New single-sheet workbook - the common case for every export below
    except the two multi-sheet ones (duplicates, diff report), which build
    their own sheets by calling write_sheet per sheet instead."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    write_sheet(ws, columns, rows, widths)
    return wb


def send_workbook(wb: openpyxl.Workbook, download_name: str) -> Response:
    """The repeated BytesIO/save/seek(0)/send_file boilerplate every export
    route used to duplicate."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=download_name, mimetype=XLSX_MIMETYPE)


# Column shape shared by the "list of current asset rows" exports (Branch
# Detail as-is; Manage Assets and User Asset History splice in their own
# extra Branch/Period column since their row shape isn't identical) - kept
# as a plain reusable constant rather than a hidden default, so each call
# site stays explicit about what it adds.
ASSET_ROW_COLUMNS: list[ColumnSpec] = [
    ("Device", "device_name"),
    ("User ID", "user_id_raw"),
    ("Full Name", "full_name"),
    ("Model", "model_device"),
    ("Serial/Service Tag", "serial_tag"),
    ("Status", "status"),
    ("Remark", "remark"),
    ("Position", "position"),
    ("Handover Date", "handover_date"),
    ("Usage Duration", lambda r: usage_duration_years(r["handover_date"])),
]


def build_asset_rows_workbook(
    rows: Iterable, sheet_title: str = "Assets", columns: Sequence[ColumnSpec] = ASSET_ROW_COLUMNS
) -> openpyxl.Workbook:
    return build_workbook(sheet_title, columns, rows)


def build_import_template_workbook(headers: Sequence[str], example_row: Sequence[Any] | None = None) -> openpyxl.Workbook:
    """Blank downloadable template for one of the import formats. `headers`
    must come from importer.py's own constants (HEADER_ALIASES's canonical
    aliases, or its required-column constants) - never a second hand-typed
    copy, so the template can't silently drift from what the importer
    actually accepts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(list(headers))
    if example_row is not None:
        ws.append(list(example_row))
    style_header_row(ws)
    return wb


# --- Duplicates workbook (moved from asset_edit.py) -------------------------

DUPLICATE_COLUMNS: list[ColumnSpec] = [
    ("Serial", lambda item: item[0]),
    ("Branch", lambda item: (item[1]["branch_eng_name"] if "branch_eng_name" in item[1].keys() else "")
     or item[1]["branch_dept"]),
    ("Device", lambda item: item[1]["device_name"]),
    ("User ID", lambda item: item[1]["user_id_raw"]),
    ("Full Name", lambda item: item[1]["full_name"]),
    ("Model", lambda item: item[1]["model_device"]),
    ("Status", lambda item: item[1]["status"]),
    ("Asset ID", lambda item: item[1]["id"]),
]


def build_duplicates_workbook(dupes: list) -> openpyxl.Workbook:
    """Shared by both places duplicates get exported: the system-wide
    Duplicate Check page (dupes from find_current_duplicate_serials) and the
    per-import Cleaning Report (dupes from find_duplicate_serials_in_batch) -
    same [{"serial", "rows"}] shape either way, just missing the joined
    branch_eng_name column in the per-import case since that query has no
    branches JOIN."""
    flat_rows = [(d["serial"], row) for d in dupes for row in d["rows"]]
    return build_workbook("Duplicates", DUPLICATE_COLUMNS, flat_rows, widths=[16, 26, 14, 14, 22, 18, 12, 10])


# --- Diff/compare workbook (moved from update_compare.py) -------------------


def build_diff_workbook(all_diffs) -> openpyxl.Workbook:
    """Shared by the on-demand /update/export download and the auto-saved
    archive copy written after every asset-report import (see
    import_data._save_diff_report) - both need the exact same sheets.

    Not built via write_sheet/build_workbook: the "Details" sheet emits a
    variable number of rows per diff (1 for each added/removed asset, but
    one row PER CHANGED FIELD for a changed asset), which doesn't reduce to
    a plain one-row-per-item column spec."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diff Summary"
    ws.append(["Branch", "Added", "Removed", "Changed", "Unchanged"])
    for d in all_diffs:
        ws.append([d.branch_label, len(d.added), len(d.removed), len(d.changed), d.unchanged_count])
    style_header_row(ws)

    detail_ws = wb.create_sheet("Details")
    # Every identifying column from the original asset report (same names/
    # order as the source file), plus the change columns at the end - so a
    # changed/added/removed row can be located and cross-checked against the
    # source spreadsheet without guessing which physical asset it was.
    header = [
        "Branch (Matched)", "Branch / Dept (as in file)", "Device Name", "User ID",
        "Full Name", "Model Device", "IP", "Serial / Service Tag", "Status", "Remark",
        "Position", "Handover Date",
        "Change Type", "Field Changed", "Old Value", "New Value",
    ]
    detail_ws.append(header)

    def _row_columns(branch_label: str, row) -> list:
        return [
            branch_label, row["branch_dept"], row["device_name"], row["user_id_raw"],
            row["full_name"], row["model_device"], row["ip"], row["serial_tag"], row["status"],
            row["remark"], row["position"], row["handover_date"],
        ]

    for d in all_diffs:
        for row in d.added:
            detail_ws.append([*_row_columns(d.branch_label, row), "ADDED", "", "", ""])
        for row in d.removed:
            detail_ws.append([*_row_columns(d.branch_label, row), "REMOVED", "", "", ""])
        for c in d.changed:
            for f in COMPARE_FIELDS:
                if f in c["diffs"]:
                    old_val, new_val = c["diffs"][f]
                    detail_ws.append(
                        [*_row_columns(d.branch_label, c["new"]), "CHANGED", FIELD_LABELS[f], old_val, new_val]
                    )
    style_header_row(detail_ws)

    return wb
